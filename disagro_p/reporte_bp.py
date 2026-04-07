from ctypes import alignment
import json
import datetime 
from datetime import date
from openpyxl.styles import Font, Alignment
from flask import (
    Blueprint,
    flash,
    g,
    redirect,
    render_template,
    request,
    session,
    url_for,
    jsonify,
)
from flask.helpers import make_response
from werkzeug.exceptions import abort
from disagro_i.auth import requiere_login, role_required
# from disagro_i.clases import ( caja as class_caja, pedido as class_pedido, tienda as class_tienda)
import os
import csv
from io import TextIOWrapper
from collections import defaultdict, OrderedDict
from copy import deepcopy
import sys
from disagro_i.clases import modelo
from disagro_i.conexion_orm import SessionLocal
from sqlalchemy.orm import Session
from sqlalchemy import asc,desc,update,func,case,cast,String,or_
import requests
import datetime
import time
import base64
import json
import csv
from sqlalchemy.exc import SQLAlchemyError
import openpyxl
import json
from sqlalchemy.sql import literal_column
from disagro_i.clases.utils import Utils
import pprint
import xlsxwriter
from io import BytesIO
from typing import List, Dict, Tuple
from urllib.parse import urlencode

bp = Blueprint('reporte_bp', __name__, url_prefix='/reporte')


def _get_planificacion_actual(db: Session, id_planificacion: str):
    """
    Devuelve la planificación solicitada. Si el usuario no es el creador,
    se permite acceder siempre que tenga rol nivel_1, nivel_2 o nivel_3.
    """
    planificacion = db.query(modelo.Planificacion).filter(modelo.Planificacion.ID == id_planificacion).first()
    if not planificacion:
        abort(404)

    es_propietario = planificacion.USUARIO == g.user.usuario
    roles = (session.get('roles') or {})
    tiene_rol_auditoria = any(roles.get(r) == 'SI' for r in ['nivel_1', 'nivel_2', 'nivel_3'])

    if not es_propietario and not tiene_rol_auditoria:
        abort(403)
    return planificacion


def _build_planificacion_context(planificacion: modelo.Planificacion) -> Dict[str, object]:
    """Assemble template context keys related to planificacion approval metadata."""
    return {
        "observacion_cierre": planificacion.OBSERVACION_CIERRE,
        "usuario_aprobacion": planificacion.USUARIO_APROBACION,
        "fecha_aprobacion": planificacion.FECHA_APROBACION,
        "planificacion_nombre": planificacion.NOMBRE,
        "planificacion_estado": planificacion.REPORTE_ESTADO,
        "planificacion_estado_interno": planificacion.ESTADO,
        "planificacion_aprobada": planificacion.REPORTE_ESTADO == "APROBADO",
        "planificacion_correlativo": planificacion.CORRELATIVO,
}


def _obtener_dimensiones_planificacion(
    db: Session,
    id_planificacion: str,
    dimension: str,
) -> List[Dict[str, str]]:
    """Return option list (value/label) constrained to the planificación scope."""

    dimension = dimension.upper()
    if dimension not in {"UBICACION", "ALMACEN"}:
        return []

    lineas = [
        linea.VALOR_FILTRO
        for linea in db.query(modelo.Planificacion_linea)
        .filter(
            modelo.Planificacion_linea.PLANIFICACION_ID == id_planificacion,
            modelo.Planificacion_linea.NOMBRE_TABLA_FILTRO == dimension,
        )
        .all()
        if linea.VALOR_FILTRO
    ]

    incluye_todas = any(valor.upper().startswith("TODAS") for valor in lineas)

    if dimension == "UBICACION":
        codigo_attr = modelo.Ubicacion.UBICACION
        descripcion_attr = modelo.Ubicacion.DESCRIPCION
        existencia_attr = modelo.Existencia.UBICACION
        tabla_modelo = modelo.Ubicacion
    else:
        codigo_attr = modelo.Almacen.ALMACEN
        descripcion_attr = modelo.Almacen.DESCRIPCION
        existencia_attr = modelo.Existencia.ALMACEN
        tabla_modelo = modelo.Almacen

    opciones: List[Dict[str, str]] = []

    if lineas and not incluye_todas:
        codigos = sorted({valor for valor in lineas})
        descripciones = dict(
            db.query(codigo_attr.label("codigo"), descripcion_attr.label("descripcion"))
            .filter(codigo_attr.in_(codigos))
            .all()
        )
        for codigo in codigos:
            label = (descripciones.get(codigo) or codigo).strip()
            opciones.append({"value": codigo, "label": label})
        return opciones

    registros_existencia = (
        db.query(
            existencia_attr.label("codigo"),
            func.max(descripcion_attr).label("descripcion"),
        )
        .outerjoin(tabla_modelo, codigo_attr == existencia_attr)
        .filter(modelo.Existencia.ID_PLANIFICACION == id_planificacion)
        .group_by(existencia_attr)
        .order_by(existencia_attr.asc())
        .all()
    )

    for registro in registros_existencia:
        codigo = registro.codigo
        if not codigo:
            continue
        descripcion = (registro.descripcion or codigo).strip()
        opciones.append({"value": codigo, "label": descripcion})

    return opciones


def build_report_catalog(id_planificacion: str) -> List[Dict[str, str]]:
    """Return the list of report and action options for the selector."""

    option_blueprints = [
        {
            "key": "diferencias",
            "label": "Inventario físico",
            "description": "Comparativo de captaciones versus planificado por ubicación.",
            "endpoint": "reporte_bp.diferencias",
            "icon": "fa-clipboard-list",
            "type": "inline",
        },
        {
            "key": "comentarios",
            "label": "Inventario con comentarios",
            "description": "Resalta captaciones con observaciones y anexa el detalle de comentarios.",
            "endpoint": "reporte_bp.diferencias_comentarios",
            "icon": "fa-comment-dots",
            "type": "inline",
        },
        {
            "key": "conteo",
            "label": "Conteo por usuario",
            "description": "Detalle de captaciones por captador con comparación contra lo planificado.",
            "endpoint": "reporte_bp.reporte_conteo",
            "icon": "fa-user-check",
            "type": "inline",
        },
        {
            "key": "consolidado",
            "label": "Consolidado cantidades/costo",
            "description": "Resumen de cantidades físicas y valoración monetaria en un solo reporte.",
            "endpoint": "reporte_bp.reporte_consolidado",
            "icon": "fa-balance-scale",
            "type": "inline",
            "required_role": "NIVEL_2",
        },
        {
            "key": "costos",
            "label": "Diferencias por costo",
            "description": "Incluye valoración monetaria de las diferencias detectadas.",
            "endpoint": "reporte_bp.diferencias_costo",
            "icon": "fa-coins",
            "type": "inline",
            "required_role": "NIVEL_2",
        },
        {
            "key": "nuevos",
            "label": "Artículos nuevos",
            "description": "Listado de artículos sin registro previo en el sistema.",
            "endpoint": "reporte_bp.nuevos_articulos",
            "icon": "fa-star",
            "type": "inline",
        },
        {
            "key": "estados",
            "label": "Reporte por estados",
            "description": "Agrupa las captaciones según su estado operativo.",
            "endpoint": "reporte_bp.reporte_estados",
            "icon": "fa-project-diagram",
            "type": "inline",
        },
        {
            "key": "galeria",
            "label": "Galería de estados",
            "description": "Visualiza las fotografías asociadas a cada captación.",
            "endpoint": "reporte_bp.reporte_imagenes",
            "icon": "fa-images",
            "type": "inline",
        },
        {
            "key": "transito",
            "label": "Artículos en tránsito",
            "description": "Inventario en movimiento entre ubicaciones.",
            "endpoint": "reporte_bp.reporte_transito",
            "icon": "fa-truck",
            "type": "inline",
        },
    ]

    catalog: List[Dict[str, str]] = []
    for option in option_blueprints:
        option_data = option.copy()
        option_data["url"] = url_for(option["endpoint"], id_planificacion=id_planificacion)
        catalog.append(option_data)

    return catalog


def _parse_multiselect_arg(args, base_key: str) -> List[str]:
    """Return normalized values for multi-select inputs supporting [] and comma formats."""
    raw_values: List[str] = []
    keys = [base_key, f"{base_key}[]"]
    for key in keys:
        if key in args:
            raw_values.extend(args.getlist(key))

    cleaned: List[str] = []
    for raw in raw_values:
        if not raw:
            continue
        parts = [fragment.strip() for fragment in raw.split(',') if fragment.strip()]
        cleaned.extend(parts if parts else [raw.strip()])
    return cleaned


def _build_conteo_filter_options(
    db: Session,
    utils: Utils,
    id_planificacion: str,
    pais: str,
    args,
    planificacion: modelo.Planificacion,
) -> Dict[str, object]:
    """Resolve available options and current selections for el reporte de conteo."""

    usuarios_planificados = utils.usuarios_planificados(db, id_planificacion, pais)
    captador_options = [
        {
            "value": usuario.usuario,
            "label": (usuario.nombre or usuario.usuario).strip() if getattr(usuario, "nombre", None) else usuario.usuario,
        }
        for usuario in usuarios_planificados
        if usuario
    ]
    captador_labels = {op["value"]: op["label"] for op in captador_options}

    plan_owner_username = getattr(planificacion, "USUARIO", None)
    if plan_owner_username and plan_owner_username not in captador_labels:
        owner = (
            db.query(modelo.Usuario)
            .filter(modelo.Usuario.usuario == plan_owner_username)
            .first()
        )
        if owner:
            label = (owner.nombre or owner.usuario).strip() if getattr(owner, "nombre", None) else owner.usuario
        else:
            label = plan_owner_username
        captador_options.append({"value": plan_owner_username, "label": label})
        captador_labels[plan_owner_username] = label

    ubicacion_options = _obtener_dimensiones_planificacion(db, id_planificacion, "UBICACION")
    ubicacion_labels = {op["value"]: op["label"] for op in ubicacion_options}

    almacen_options = _obtener_dimensiones_planificacion(db, id_planificacion, "ALMACEN")
    almacen_labels = {op["value"]: op["label"] for op in almacen_options}

    raw_captadores = _parse_multiselect_arg(args, 'captadores')
    selected_captadores: List[str] = []
    include_sin_captador = False
    for value in raw_captadores:
        if value.upper() == 'SIN_CAPTADOR':
            include_sin_captador = True
            continue
        if value in captador_labels:
            selected_captadores.append(value)

    if not selected_captadores:
        selected_captadores = [op["value"] for op in captador_options]

    raw_ubicaciones = _parse_multiselect_arg(args, 'ubicaciones')
    selected_ubicaciones = [
        value for value in raw_ubicaciones
        if value in ubicacion_labels and value.upper() != 'TODAS'
    ]

    raw_almacenes = _parse_multiselect_arg(args, 'almacenes')
    selected_almacenes = [
        value for value in raw_almacenes
        if value in almacen_labels and value.upper() != 'TODAS'
    ]

    captadores_display: List[str] = []
    if selected_captadores and len(selected_captadores) != len(captador_options):
        captadores_display = list(selected_captadores)
    if include_sin_captador:
        captadores_display.append('SIN_CAPTADOR')

    filters = {
        "options": {
            "captadores": captador_options,
            "ubicaciones": ubicacion_options,
            "almacenes": almacen_options,
        },
        "labels": {
            "captadores": captador_labels,
            "ubicaciones": ubicacion_labels,
            "almacenes": almacen_labels,
            "sin_captador": "Sin captador",
        },
        "selected": {
            "captadores": selected_captadores,
            "captadores_display": captadores_display,
            "include_sin": include_sin_captador,
            "captadores_all": len(selected_captadores) == len(captador_options) and not include_sin_captador,
            "ubicaciones": selected_ubicaciones,
            "ubicaciones_all": len(selected_ubicaciones) == 0,
            "almacenes": selected_almacenes,
            "almacenes_all": len(selected_almacenes) == 0,
        },
        "criteria": {
            "captadores": selected_captadores,
            "include_sin": include_sin_captador,
            "ubicaciones": selected_ubicaciones,
            "almacenes": selected_almacenes,
        },
    }
    return filters


def _obtener_observaciones_captaciones(
    db: Session,
    id_planificacion: str,
) -> Tuple[Dict[Tuple[object, ...], List[Dict[str, object]]], List[Dict[str, object]]]:
    """Return mapping and flat list for captación comments within the plan."""

    registros = (
        db.query(
            modelo.CaptacionFisica.ID,
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            modelo.CaptacionFisica.FECHA_EXPIRACION,
            modelo.CaptacionFisica.SERIE,
            modelo.CaptacionFisica.MODELO,
            modelo.CaptacionFisica.ETIQUETA,
            modelo.CaptacionFisica.OBSERVACION,
            modelo.CaptacionFisica.USUARIO,
            modelo.CaptacionFisica.FECHA,
            modelo.CaptacionFisica.CANTIDAD,
        )
        .filter(
            modelo.CaptacionFisica.ID_PLANIFICACION == id_planificacion,
            modelo.CaptacionFisica.OBSERVACION.isnot(None),
            func.trim(modelo.CaptacionFisica.OBSERVACION) != '',
            or_(
                modelo.CaptacionFisica.ETIQUETA.is_(None),
                modelo.CaptacionFisica.ETIQUETA != 'EN_TRANSITO',
            ),
        )
        .order_by(
            modelo.CaptacionFisica.UBICACION.asc(),
            modelo.CaptacionFisica.ARTICULO.asc(),
            modelo.CaptacionFisica.ALMACEN.asc(),
            modelo.CaptacionFisica.LOTE.asc(),
            modelo.CaptacionFisica.SERIE.asc(),
            modelo.CaptacionFisica.FECHA.asc(),
        )
        .all()
    )

    comentarios_por_clave: Dict[Tuple[object, ...], List[Dict[str, object]]] = defaultdict(list)
    comentarios_lista: List[Dict[str, object]] = []

    for registro in registros:
        observacion = (registro.OBSERVACION or '').strip()
        if not observacion:
            continue

        clave = (
            registro.UBICACION,
            registro.ARTICULO,
            registro.ALMACEN,
            registro.LOTE,
            registro.FECHA_EXPIRACION,
            registro.SERIE,
            registro.MODELO,
            registro.ETIQUETA,
        )

        comentario = {
            "ID": registro.ID,
            "UBICACION": registro.UBICACION,
            "ALMACEN": registro.ALMACEN,
            "ARTICULO": registro.ARTICULO,
            "DESCRIPCION": registro.DESCRIPCION,
            "LOTE": registro.LOTE,
            "FECHA_EXPIRACION": registro.FECHA_EXPIRACION,
            "SERIE": registro.SERIE,
            "MODELO": registro.MODELO,
            "OBSERVACION": observacion,
            "USUARIO": registro.USUARIO,
            "FECHA": registro.FECHA,
            "CANTIDAD": registro.CANTIDAD,
            "ETIQUETA": registro.ETIQUETA,
        }
        comentarios_por_clave[clave].append(comentario)
        comentarios_lista.append(comentario)

    return comentarios_por_clave, comentarios_lista


def _build_existencias_dict(existencias: List[Dict[str, object]]) -> Dict[Tuple[object, ...], Dict[str, object]]:
    """Index existencias by their dimensional keys for quick lookups."""
    index: Dict[Tuple[object, ...], Dict[str, object]] = {}
    for existencia in existencias:
        key = (
            existencia.get("ARTICULO"),
            existencia.get("UBICACION"),
            existencia.get("ALMACEN"),
            existencia.get("LOTE"),
            existencia.get("FECHA_EXPIRACION"),
        )
        index[key] = existencia
    return index


def obtener_reporte_conteo(db: Session, id_planificacion: str, pais: str, criterios: Dict[str, object]) -> Dict[str, object]:
    """Collect raw rows and totals for the conteo report respecting filters."""

    utils = Utils()
    filtros_base = list(utils.obtener_filtros(db, id_planificacion, pais))

    ubicaciones = criterios.get("ubicaciones", []) or []
    almacenes = criterios.get("almacenes", []) or []

    if ubicaciones:
        filtros_base.append(modelo.Existencia.UBICACION.in_(ubicaciones))
    if almacenes:
        filtros_base.append(modelo.Existencia.ALMACEN.in_(almacenes))

    existencias = [dict(row._mapping) for row in utils.obtener_existencias_planificadas(db, filtros_base, id_planificacion)]
    existencias_dict = _build_existencias_dict(existencias)

    captadores = criterios.get("captadores", []) or []
    include_sin_captador = criterios.get("include_sin", False)

    captacion_rows: List[Dict[str, object]] = []
    if captadores:
        consulta_captaciones = db.query(
            modelo.CaptacionFisica.USUARIO.label("USUARIO"),
            modelo.CaptacionFisica.ARTICULO.label("ARTICULO"),
            modelo.CaptacionFisica.DESCRIPCION.label("DESCRIPCION"),
            modelo.CaptacionFisica.UBICACION.label("UBICACION"),
            modelo.CaptacionFisica.ALMACEN.label("ALMACEN"),
            modelo.CaptacionFisica.LOTE.label("LOTE"),
            modelo.CaptacionFisica.FECHA_EXPIRACION.label("FECHA_EXPIRACION"),
            func.max(modelo.CaptacionFisica.FECHA).label("FECHA_CAPTURA"),
            func.round(func.sum(modelo.CaptacionFisica.CANTIDAD), 3).label("CANTIDAD_CAPTACION"),
        ).filter(
            modelo.CaptacionFisica.ID_PLANIFICACION == id_planificacion,
            modelo.CaptacionFisica.USUARIO.in_(captadores),
            or_(
                modelo.CaptacionFisica.ETIQUETA.is_(None),
                modelo.CaptacionFisica.ETIQUETA != 'EN_TRANSITO'
            )
        )

        if ubicaciones:
            consulta_captaciones = consulta_captaciones.filter(modelo.CaptacionFisica.UBICACION.in_(ubicaciones))
        if almacenes:
            consulta_captaciones = consulta_captaciones.filter(modelo.CaptacionFisica.ALMACEN.in_(almacenes))

        consulta_captaciones = consulta_captaciones.group_by(
            modelo.CaptacionFisica.USUARIO,
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            modelo.CaptacionFisica.FECHA_EXPIRACION,
        ).order_by(
            modelo.CaptacionFisica.UBICACION.asc(),
            modelo.CaptacionFisica.ALMACEN.asc(),
            modelo.CaptacionFisica.USUARIO.asc(),
            modelo.CaptacionFisica.ARTICULO.asc(),
        )

        captacion_rows = [dict(row._mapping) for row in consulta_captaciones.all()]

    resultados: List[Dict[str, object]] = []
    claves_captadas: Dict[Tuple[object, ...], bool] = {}

    for row in captacion_rows:
        key = (
            row.get("ARTICULO"),
            row.get("UBICACION"),
            row.get("ALMACEN"),
            row.get("LOTE"),
            row.get("FECHA_EXPIRACION"),
        )
        existencia = existencias_dict.get(key)
        cantidad_existencia = float(existencia.get("cantidad_existencia", 0)) if existencia else 0.0
        cantidad_captacion = float(row.get("CANTIDAD_CAPTACION", 0) or 0)
        diferencia = round(cantidad_captacion - cantidad_existencia, 3)
        fecha_exp = row.get("FECHA_EXPIRACION")
        fecha_cap = row.get("FECHA_CAPTURA")
        fecha_exp_str = None
        fecha_cap_str = None
        if isinstance(fecha_exp, datetime.datetime):
            fecha_exp_str = fecha_exp.strftime('%d/%m/%Y')
        elif isinstance(fecha_exp, datetime.date):
            fecha_exp_str = fecha_exp.strftime('%d/%m/%Y')
        if isinstance(fecha_cap, datetime.datetime):
            fecha_cap_str = fecha_cap.strftime('%d/%m/%Y %H:%M')
        elif isinstance(fecha_cap, datetime.date):
            fecha_cap_str = fecha_cap.strftime('%d/%m/%Y')

        resultados.append({
            "ARTICULO": row.get("ARTICULO"),
            "DESCRIPCION": row.get("DESCRIPCION") or (existencia.get("DESCRIPCION") if existencia else None),
            "UBICACION": row.get("UBICACION"),
            "ALMACEN": row.get("ALMACEN"),
            "LOTE": row.get("LOTE"),
            "FECHA_EXPIRACION": row.get("FECHA_EXPIRACION"),
            "FECHA_EXPIRACION_STR": fecha_exp_str,
            "USUARIO": row.get("USUARIO"),
            "CANTIDAD_EXISTENCIA": round(cantidad_existencia, 3),
            "CANTIDAD_CAPTACION": round(cantidad_captacion, 3),
            "DIFERENCIA": diferencia,
            "FECHA_CAPTURA": row.get("FECHA_CAPTURA"),
            "FECHA_CAPTURA_STR": fecha_cap_str,
        })
        claves_captadas[key] = True

    if include_sin_captador:
        for key, existencia in existencias_dict.items():
            if key in claves_captadas:
                continue
            cantidad_existencia = float(existencia.get("cantidad_existencia", 0) or 0)
            resultados.append({
                "ARTICULO": existencia.get("ARTICULO"),
                "DESCRIPCION": existencia.get("DESCRIPCION"),
                "UBICACION": existencia.get("UBICACION"),
                "ALMACEN": existencia.get("ALMACEN"),
                "LOTE": existencia.get("LOTE"),
                "FECHA_EXPIRACION": existencia.get("FECHA_EXPIRACION"),
                "FECHA_EXPIRACION_STR": None,
                "USUARIO": 'SIN_CAPTADOR',
                "CANTIDAD_EXISTENCIA": round(cantidad_existencia, 3),
                "CANTIDAD_CAPTACION": 0.0,
                "DIFERENCIA": round(-cantidad_existencia, 3),
                "FECHA_CAPTURA": None,
                "FECHA_CAPTURA_STR": None,
            })

    resultados.sort(key=lambda r: (
        r.get("UBICACION") or "",
        r.get("ALMACEN") or "",
        r.get("USUARIO") or "",
        r.get("ARTICULO") or "",
        r.get("LOTE") or "",
        r.get("FECHA_EXPIRACION") or datetime.datetime.min,
    ))

    total_plan = round(sum(row.get("CANTIDAD_EXISTENCIA", 0) or 0 for row in resultados), 3)
    total_captado = round(sum(row.get("CANTIDAD_CAPTACION", 0) or 0 for row in resultados), 3)
    total_diferencia = round(sum(row.get("DIFERENCIA", 0) or 0 for row in resultados), 3)

    resumen_por_usuario: Dict[str, Dict[str, float]] = {}
    for row in resultados:
        usuario = row.get("USUARIO") or "SIN_CAPTADOR"
        data = resumen_por_usuario.setdefault(usuario, {"plan": 0.0, "captado": 0.0, "diferencia": 0.0})
        data["plan"] += row.get("CANTIDAD_EXISTENCIA", 0) or 0
        data["captado"] += row.get("CANTIDAD_CAPTACION", 0) or 0
        data["diferencia"] += row.get("DIFERENCIA", 0) or 0

    for resumen in resumen_por_usuario.values():
        resumen["plan"] = round(resumen["plan"], 3)
        resumen["captado"] = round(resumen["captado"], 3)
        resumen["diferencia"] = round(resumen["diferencia"], 3)

    resumen_ordenado = dict(sorted(resumen_por_usuario.items(), key=lambda item: item[0]))

    return {
        "rows": resultados,
        "totales": {
            "plan": total_plan,
            "captado": total_captado,
            "diferencia": total_diferencia,
        },
        "totales_por_usuario": resumen_ordenado,
    }


def agrupar_conteo_por_ubicacion(conteo_rows: List[Dict[str, object]]) -> Dict[str, object]:
    """Construye la estructura presentada en la tabla estilo diferencias."""

    agrupado: Dict[str, Dict[str, Dict[str, object]]] = defaultdict(
        lambda: defaultdict(lambda: {"DESCRIPCION": "", "CAPTACIONES": []})
    )

    for fila in conteo_rows:
        ubicacion = fila.get("UBICACION") or "-"
        articulo = fila.get("ARTICULO") or "-"

        articulo_entry = agrupado[ubicacion][articulo]
        articulo_entry["DESCRIPCION"] = fila.get("DESCRIPCION") or ""
        articulo_entry["CAPTACIONES"].append({
            "ALMACEN": fila.get("ALMACEN") or "",
            "LOTE": fila.get("LOTE") or "",
            "SERIE": fila.get("SERIE") or "",
            "MODELO": fila.get("MODELO") or "",
            "FECHA_EXPIRACION": fila.get("FECHA_EXPIRACION"),
            "FISICO": round(float(fila.get("CANTIDAD_CAPTACION", 0) or 0), 3),
            "SISTEMA": round(float(fila.get("CANTIDAD_EXISTENCIA", 0) or 0), 3),
            "DIFERENCIA": round(float(fila.get("DIFERENCIA", 0) or 0), 3),
            "USUARIO": fila.get("USUARIO") or "SIN_CAPTADOR",
            "FECHA_CAPTURA": fila.get("FECHA_CAPTURA"),
        })

    presentacion = {"captaciones_por_ubicacion": []}

    diferencia_total = 0.0
    for ubicacion, articulos in agrupado.items():
        ubicacion_data = {
            "UBICACION": ubicacion,
            "ARTICULOS": [],
            "TOTAL_FISICO": 0.0,
            "TOTAL_SISTEMA": 0.0,
            "TOTAL_DIFERENCIA": 0.0,
        }

        for articulo, detalles in articulos.items():
            captaciones = detalles.get("CAPTACIONES", [])
            total_fisico = round(sum(c["FISICO"] for c in captaciones), 3)
            total_sistema = round(sum(c["SISTEMA"] for c in captaciones), 3)
            total_diferencia = round(sum(float(c["DIFERENCIA"]) for c in captaciones), 3)

            articulo_data = {
                "ARTICULO": articulo,
                "DESCRIPCION": detalles.get("DESCRIPCION", ""),
                "CAPTACIONES": captaciones,
                "CAPTACIONES_NUEVAS": [],
                "TOTAL_FISICO": total_fisico,
                "TOTAL_SISTEMA": total_sistema,
                "TOTAL_DIFERENCIA": total_diferencia,
            }

            ubicacion_data["ARTICULOS"].append(articulo_data)
            ubicacion_data["TOTAL_FISICO"] = round(ubicacion_data["TOTAL_FISICO"] + total_fisico, 3)
            ubicacion_data["TOTAL_SISTEMA"] = round(ubicacion_data["TOTAL_SISTEMA"] + total_sistema, 3)
            ubicacion_data["TOTAL_DIFERENCIA"] = round(ubicacion_data["TOTAL_DIFERENCIA"] + total_diferencia, 3)

        diferencia_total += ubicacion_data["TOTAL_DIFERENCIA"]
        presentacion["captaciones_por_ubicacion"].append(ubicacion_data)

    presentacion["DIFERENCIA_TOTAL"] = round(diferencia_total, 3)
    return presentacion


def obtener_diccionario_consolidado(resultados: List[Dict[str, object]]) -> Dict[str, object]:
    """Combina cantidades físicas y costos en la estructura estilo diferencias."""

    ubicaciones: "OrderedDict[str, OrderedDict[str, Dict[str, object]]]" = OrderedDict()

    for row in resultados:
        ubicacion = row.get("UBICACION") or "-"
        articulo = row.get("ARTICULO") or "-"

        ubicacion_dict = ubicaciones.setdefault(ubicacion, OrderedDict())
        articulo_dict = ubicacion_dict.setdefault(
            articulo,
            {
                "DESCRIPCION": row.get("DESCRIPCION") or "",
                "CAPTACIONES": [],
            },
        )

        costo_unitario = float(row.get("costo_existencia") or 0)
        fisico = float(row.get("cantidad_captacion", 0) or 0)
        sistema = float(row.get("cantidad_existencia", 0) or 0)
        diferencia = float(row.get("diferencia", 0) or 0)

        fecha_exp = row.get("FECHA_EXPIRACION")
        if isinstance(fecha_exp, datetime.datetime):
            fecha_exp_str = fecha_exp.strftime('%d/%m/%Y')
        elif isinstance(fecha_exp, datetime.date):
            fecha_exp_str = fecha_exp.strftime('%d/%m/%Y')
        else:
            fecha_exp_str = row.get("FECHA_EXPIRACION_STR")

        fecha_cap = row.get("FECHA_CAPTURA")
        if isinstance(fecha_cap, datetime.datetime):
            fecha_cap_str = fecha_cap.strftime('%d/%m/%Y %H:%M')
        elif isinstance(fecha_cap, datetime.date):
            fecha_cap_str = fecha_cap.strftime('%d/%m/%Y')
        else:
            fecha_cap_str = row.get("FECHA_CAPTURA_STR")

        fisico_costo = round(fisico * costo_unitario, 3)
        sistema_costo = round(sistema * costo_unitario, 3)
        diferencia_costo = round(diferencia * costo_unitario, 3)

        articulo_dict["CAPTACIONES"].append(
            {
                "ALMACEN": row.get("ALMACEN") or "",
                "LOTE": row.get("LOTE") or "",
                "SERIE": row.get("SERIE") or "",
                "MODELO": row.get("MODELO") or "",
                "FECHA_EXPIRACION": fecha_exp,
                "FECHA_EXPIRACION_STR": fecha_exp_str,
                "FISICO": round(fisico, 3),
                "SISTEMA": round(sistema, 3),
                "DIFERENCIA": round(diferencia, 3),
                "FISICO_COSTO": fisico_costo,
                "SISTEMA_COSTO": sistema_costo,
                "DIFERENCIA_COSTO": diferencia_costo,
                "USUARIO": row.get("USUARIO") or "SIN_CAPTADOR",
                "FECHA_CAPTURA": fecha_cap,
                "FECHA_CAPTURA_STR": fecha_cap_str,
            }
        )

    datos = {"captaciones_por_ubicacion": []}
    diferencia_total = 0.0
    diferencia_total_costo = 0.0

    for ubicacion, articulos in ubicaciones.items():
        ubicacion_data = {
            "UBICACION": ubicacion,
            "ARTICULOS": [],
            "TOTAL_FISICO": 0.0,
            "TOTAL_SISTEMA": 0.0,
            "TOTAL_DIFERENCIA": 0.0,
            "TOTAL_FISICO_COSTO": 0.0,
            "TOTAL_SISTEMA_COSTO": 0.0,
            "TOTAL_DIFERENCIA_COSTO": 0.0,
        }

        for articulo, detalles in articulos.items():
            captaciones = detalles.get("CAPTACIONES", [])
            total_fisico = round(sum(c["FISICO"] for c in captaciones), 3)
            total_sistema = round(sum(c["SISTEMA"] for c in captaciones), 3)
            total_diferencia_art = round(sum(c["DIFERENCIA"] for c in captaciones), 3)
            total_fisico_costo = round(sum(c["FISICO_COSTO"] for c in captaciones), 3)
            total_sistema_costo = round(sum(c["SISTEMA_COSTO"] for c in captaciones), 3)
            total_diferencia_costo_art = round(sum(c["DIFERENCIA_COSTO"] for c in captaciones), 3)

            articulo_data = {
                "ARTICULO": articulo,
                "DESCRIPCION": detalles.get("DESCRIPCION", ""),
                "CAPTACIONES": captaciones,
                "CAPTACIONES_NUEVAS": [],
                "TOTAL_FISICO": total_fisico,
                "TOTAL_SISTEMA": total_sistema,
                "TOTAL_DIFERENCIA": total_diferencia_art,
                "TOTAL_FISICO_COSTO": total_fisico_costo,
                "TOTAL_SISTEMA_COSTO": total_sistema_costo,
                "TOTAL_DIFERENCIA_COSTO": total_diferencia_costo_art,
            }

            ubicacion_data["ARTICULOS"].append(articulo_data)
            ubicacion_data["TOTAL_FISICO"] = round(ubicacion_data["TOTAL_FISICO"] + total_fisico, 3)
            ubicacion_data["TOTAL_SISTEMA"] = round(ubicacion_data["TOTAL_SISTEMA"] + total_sistema, 3)
            ubicacion_data["TOTAL_DIFERENCIA"] = round(ubicacion_data["TOTAL_DIFERENCIA"] + total_diferencia_art, 3)
            ubicacion_data["TOTAL_FISICO_COSTO"] = round(ubicacion_data["TOTAL_FISICO_COSTO"] + total_fisico_costo, 3)
            ubicacion_data["TOTAL_SISTEMA_COSTO"] = round(ubicacion_data["TOTAL_SISTEMA_COSTO"] + total_sistema_costo, 3)
            ubicacion_data["TOTAL_DIFERENCIA_COSTO"] = round(
                ubicacion_data["TOTAL_DIFERENCIA_COSTO"] + total_diferencia_costo_art, 2
            )

        diferencia_total += ubicacion_data["TOTAL_DIFERENCIA"]
        diferencia_total_costo += ubicacion_data["TOTAL_DIFERENCIA_COSTO"]
        datos["captaciones_por_ubicacion"].append(ubicacion_data)

    datos["DIFERENCIA_TOTAL"] = round(diferencia_total, 3)
    datos["DIFERENCIA_TOTAL_COSTO"] = round(diferencia_total_costo, 3)
    return datos


@bp.route("/diferencias/<id_planificacion>",methods=['GET'])
@requiere_login
@role_required(['nivel_1','nivel_2','nivel_3','nivel_4'])
def diferencias(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db,g)
        resultados = []
        costo = False

        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA

        if planificacion.ESTADO == "EXISTENCIAS_CARGADAS":
            flash("La planificación no ha sido completada, aún no se ha planificado.")
            return redirect(url_for('planificacion_bp.planificaciones'))
        
        #obtener_diferencias se encarga de obtener las diferencias entre las captaciones fisicas y las existencias planificadas
        resultados = obtener_diferencias(db,id_planificacion,pais)
        #obtener_diccionario se encarga de transformar los resultados en un diccionario
        diccionario = obtener_diccionario(resultados)
        datos_diferencias = captaciones_por_ubicacion(diccionario)
        seccion_de_planificacion = True
        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": date,
            "diferencias": datos_diferencias,
            "id_planificacion": id_planificacion,
            "seccion_de_planificacion": seccion_de_planificacion,
            "costo": costo,
            "report_catalog": report_catalog,
            "selected_report": "diferencias",
            "solo_nuevos": False,
            "sin_datos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()


@bp.route("/diferencias/comentarios/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def diferencias_comentarios(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)

        planificacion = _get_planificacion_actual(db, id_planificacion)
        now = planificacion.FECHA

        if planificacion.ESTADO == "EXISTENCIAS_CARGADAS":
            flash("La planificación no ha sido completada, aún no se ha planificado.")
            return redirect(url_for('planificacion_bp.planificaciones'))

        resultados = obtener_diferencias(db, id_planificacion, pais)
        comentarios_por_clave, comentarios_detalle = _obtener_observaciones_captaciones(db, id_planificacion)
        diccionario = obtener_diccionario(resultados, comentarios_por_clave)
        datos_diferencias = captaciones_por_ubicacion(diccionario)

        comentarios_ordenados = sorted(
            comentarios_detalle,
            key=lambda item: (
                item.get("UBICACION") or "",
                item.get("ARTICULO") or "",
                item.get("ALMACEN") or "",
                item.get("LOTE") or "",
                item.get("SERIE") or "",
                item.get("FECHA") or datetime.datetime.min,
            ),
        )

        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": now,
            "diferencias": datos_diferencias,
            "id_planificacion": id_planificacion,
            "seccion_de_planificacion": True,
            "costo": False,
            "report_catalog": report_catalog,
            "selected_report": "comentarios",
            "solo_nuevos": False,
            "sin_datos": len(datos_diferencias.get("captaciones_por_ubicacion", [])) == 0,
            "comentarios_resumen": comentarios_ordenados,
            "comentarios_totales": len(comentarios_ordenados),
            "tiene_comentarios": len(comentarios_ordenados) > 0,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()


@bp.route("/conteo/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def reporte_conteo(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)

        planificacion = _get_planificacion_actual(db, id_planificacion)
        now = planificacion.FECHA
        filtros = _build_conteo_filter_options(db, utils, id_planificacion, pais, request.args, planificacion)
        datos_conteo = obtener_reporte_conteo(db, id_planificacion, pais, filtros["criteria"])
        conteo_presentacion = agrupar_conteo_por_ubicacion(datos_conteo["rows"])

        total_registros = len(datos_conteo["rows"])

        if request.args.get('format') == 'json':
            def serializar(row: Dict[str, object]) -> Dict[str, object]:
                data = row.copy()
                fecha_exp = data.get("FECHA_EXPIRACION")
                if isinstance(fecha_exp, (datetime.datetime, datetime.date)):
                    data["FECHA_EXPIRACION"] = fecha_exp.isoformat()
                fecha_cap = data.get("FECHA_CAPTURA")
                if isinstance(fecha_cap, (datetime.datetime, datetime.date)):
                    data["FECHA_CAPTURA"] = fecha_cap.isoformat()
                return data

            return jsonify({
                "rows": [serializar(row) for row in datos_conteo["rows"]],
                "totales": datos_conteo["totales"],
                "totales_por_usuario": datos_conteo["totales_por_usuario"],
                "total_registros": total_registros,
            })

        export_params = []
        for value in filtros["selected"]["captadores"]:
            export_params.append(('captadores', value))
        if filtros["selected"].get("include_sin"):
            export_params.append(('captadores', 'SIN_CAPTADOR'))
        for value in filtros["selected"]["ubicaciones"]:
            export_params.append(('ubicaciones', value))
        for value in filtros["selected"]["almacenes"]:
            export_params.append(('almacenes', value))
        export_query = urlencode(export_params)
        export_url = url_for('reporte_bp.conteo_xls', id_planificacion=id_planificacion)
        if export_query:
            export_url = f"{export_url}?{export_query}"

        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": now,
            "id_planificacion": id_planificacion,
            "report_catalog": report_catalog,
            "selected_report": "conteo",
            "conteo_rows": datos_conteo["rows"],
            "conteo_totales": datos_conteo["totales"],
            "conteo_totales_por_usuario": datos_conteo["totales_por_usuario"],
            "conteo_filters": filtros,
            "conteo_export_url": export_url,
            "conteo_total_registros": total_registros,
            "conteo_presentacion": conteo_presentacion,
            "sin_datos": total_registros == 0,
            "seccion_de_planificacion": False,
            "costo": False,
            "solo_nuevos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/conteo.html", **context)
    finally:
        db.close()


@bp.route("/consolidado/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_2'])
def reporte_consolidado(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)

        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA

        if planificacion.ESTADO == "EXISTENCIAS_CARGADAS":
            flash("La planificación no ha sido completada, aún no se ha planificado.")
            return redirect(url_for('planificacion_bp.planificaciones'))

        resultados = obtener_diferencias(db, id_planificacion, pais)
        datos_consolidado = obtener_diccionario_consolidado(resultados)
        sin_datos = len(datos_consolidado.get("captaciones_por_ubicacion", [])) == 0

        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": date,
            "diferencias": datos_consolidado,
            "id_planificacion": id_planificacion,
            "seccion_de_planificacion": False,
            "costo": False,
            "report_catalog": report_catalog,
            "selected_report": "consolidado",
            "solo_nuevos": False,
            "sin_datos": sin_datos,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()


@bp.route("/diferencias/costos/<id_planificacion>",methods=['GET'])
@requiere_login
@role_required(['nivel_2'])
def diferencias_costo(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db,g)
        resultados = []
        costo = True

        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA

        if planificacion.ESTADO == "EXISTENCIAS_CARGADAS":
            flash("La planificación no ha sido completada, aún no se ha planificado.")
            return redirect(url_for('planificacion_bp.planificaciones'))
        
        resultados = obtener_diferencias(db,id_planificacion,pais)
        diccionario = obtener_diccionario_costo(resultados)
        datos_diferencias = captaciones_por_ubicacion(diccionario)
        seccion_de_planificacion = False
        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": date,
            "diferencias": datos_diferencias,
            "id_planificacion": id_planificacion,
            "seccion_de_planificacion": seccion_de_planificacion,
            "costo": costo,
            "report_catalog": report_catalog,
            "selected_report": "costos",
            "solo_nuevos": False,
            "sin_datos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()


@bp.route("/nuevos/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1','nivel_2','nivel_3','nivel_4'])
def nuevos_articulos(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)

        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA
        if planificacion.ESTADO == "EXISTENCIAS_CARGADAS":
            flash("La planificación no ha sido completada, aún no se ha planificado.")
            return redirect(url_for('planificacion_bp.planificaciones'))

        resultados = obtener_diferencias(db, id_planificacion, pais)
        diccionario = obtener_diccionario(resultados)
        diccionario_nuevos = filtrar_diccionario_nuevos(diccionario)
        datos_nuevos = captaciones_por_ubicacion(diccionario_nuevos)
        sin_datos = len(datos_nuevos.get("captaciones_por_ubicacion", [])) == 0

        report_catalog = build_report_catalog(id_planificacion)

        context = {
            "usuario": g.user.usuario,
            "date": date,
            "diferencias": datos_nuevos,
            "id_planificacion": id_planificacion,
            "seccion_de_planificacion": False,
            "costo": False,
            "solo_nuevos": True,
            "sin_datos": sin_datos,
            "report_catalog": report_catalog,
            "selected_report": "nuevos",
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()

@bp.route("/diferencias/archivadas/<id_planificacion>",methods=['GET'])
@requiere_login
@role_required(['nivel_1','nivel_2','nivel_3','nivel_4'])
def diferencias_archivadas(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db,g)
        resultados = []
        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA
        resultados = obtener_diferencias(db,id_planificacion,pais)
        diccionario = obtener_diccionario(resultados)
        datos_diferencias = captaciones_por_ubicacion(diccionario)
        seccion_de_planificacion = False
        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": date,
            "diferencias": datos_diferencias,
            "id_planificacion": id_planificacion,
            "seccion_de_planificacion": seccion_de_planificacion,
            "report_catalog": report_catalog,
            "selected_report": "archivadas",
            "solo_nuevos": False,
            "sin_datos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()

@bp.route("/diferencias/xls/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1','nivel_2','nivel_3','nivel_4'])
def diferencias_xls(id_planificacion):
    try:
        db: Session = request.db
        date = datetime.datetime.now()
        utils = Utils()
        pais = utils.obtener_pais(db, g)
        _get_planificacion_actual(db, id_planificacion)
        resultados = obtener_diferencias(db, id_planificacion, pais)
        diccionario = obtener_diccionario(resultados)
        datos_diferencias = captaciones_por_ubicacion(diccionario)

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Diferencias")

        # Definir formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'align': 'center'})
        cell_format = workbook.add_format({'border': 1, 'align': 'center'})
        re_conteo_format = workbook.add_format({
            'border': 1, 'align': 'center', 'bg_color': '#565cb5', 'font_color': 'white'})
        nuevo_format = workbook.add_format({
            'border': 1, 'align': 'center', 'bg_color': 'red', 'font_color': 'white'})
        title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})

        row_index = 0
        col_index = 0

        # Encabezado principal
        worksheet.merge_range(row_index, col_index, row_index, col_index + 7, "REPORTE DE INVENTARIO FÍSICO", title_format)
        row_index += 2

        # Escribir fecha
        worksheet.write(row_index, 6, "FECHA:", header_format)
        worksheet.write(row_index, 7, date.strftime('%d/%m/%Y'), cell_format)
        row_index += 2

        # Recorremos cada ubicación
        for ubicacion in datos_diferencias["captaciones_por_ubicacion"]:
            # Fila de ubicación
            worksheet.merge_range(row_index, 0, row_index, 9, f"UBICACIÓN: {ubicacion['UBICACION']}", header_format)
            row_index += 1

            # Encabezado de columnas
            headers = [
                "ARTICULO",
                "DESCRIPCION",
                "ALMACEN",
                "LOTE",
                "SERIE",
                "MODELO",
                "F. EXPIRACIÓN",
                "FISICO",
                "SISTEMA",
                "DIFERENCIA",
            ]
            for col, header in enumerate(headers):
                worksheet.write(row_index, col, header, header_format)
            row_index += 1

            # Recorremos los artículos dentro de la ubicación
            for art in ubicacion["ARTICULOS"]:
                # Fila de artículo
                worksheet.write(row_index, 0, art["ARTICULO"], cell_format)
                worksheet.write(row_index, 1, art["DESCRIPCION"], cell_format)
                for col in range(2, 10):
                    worksheet.write(row_index, col, "", cell_format)
                row_index += 1

                # Recorremos las captaciones
                for captacion in art.get("CAPTACIONES", []):
                    fmt = re_conteo_format if captacion.get("RECONTEO") == "SI" else cell_format
                    worksheet.write(row_index, 2, captacion["ALMACEN"], fmt)
                    worksheet.write(row_index, 3, captacion["LOTE"] or "", fmt)
                    worksheet.write(row_index, 4, captacion.get("SERIE") or "", fmt)
                    worksheet.write(row_index, 5, captacion.get("MODELO") or "", fmt)
                    fecha_expiracion = (
                            captacion["FECHA_EXPIRACION"].strftime('%d/%m/%Y') 
                            if isinstance(captacion["FECHA_EXPIRACION"], datetime.datetime) 
                            else captacion["FECHA_EXPIRACION"] or ""
                        )
                    worksheet.write_string(row_index, 6, fecha_expiracion, fmt) 
                    worksheet.write(row_index, 7, captacion["FISICO"], fmt)
                    worksheet.write(row_index, 8, captacion["SISTEMA"], fmt)
                    worksheet.write(row_index, 9, captacion["DIFERENCIA"], fmt)
                    row_index += 1

                # Si hay captaciones nuevas
                if art.get("CAPTACIONES_NUEVAS"):
                    worksheet.write(row_index, 1, "Nuevo", nuevo_format)
                    row_index += 1
                    for captacion in art["CAPTACIONES_NUEVAS"]:
                        fmt = re_conteo_format if captacion.get("RECONTEO") == "SI" else cell_format
                        worksheet.write(row_index, 2, captacion["ALMACEN"], fmt)
                        worksheet.write(row_index, 3, captacion["LOTE"] or "", fmt)
                        worksheet.write(row_index, 4, captacion.get("SERIE") or "", fmt)
                        worksheet.write(row_index, 5, captacion.get("MODELO") or "", fmt)
                        fecha_expiracion = (
                            captacion["FECHA_EXPIRACION"].strftime('%d/%m/%Y') 
                            if isinstance(captacion["FECHA_EXPIRACION"], datetime.datetime) 
                            else captacion["FECHA_EXPIRACION"] or ""
                        )
                        worksheet.write_string(row_index, 6, fecha_expiracion, fmt) 
                        worksheet.write(row_index, 7, captacion["FISICO"], fmt)
                        worksheet.write(row_index, 8, captacion["SISTEMA"], fmt)
                        worksheet.write(row_index, 9, captacion["DIFERENCIA"], fmt)
                        row_index += 1

                # Fila total por artículo
                worksheet.write(row_index, 0, "", cell_format)
                worksheet.write(row_index, 1, "", cell_format)
                worksheet.write(row_index, 2, "TOTAL", header_format)
                worksheet.write(row_index, 7, art["TOTAL_FISICO"], header_format)
                worksheet.write(row_index, 8, art["TOTAL_SISTEMA"], header_format)
                worksheet.write(row_index, 9, art["TOTAL_DIFERENCIA"], header_format)
                row_index += 2

            # Fila total por ubicación
            worksheet.write(row_index, 0, "TOTAL", header_format)
            worksheet.write(row_index, 7, ubicacion["TOTAL_FISICO"], header_format)
            worksheet.write(row_index, 8, ubicacion["TOTAL_SISTEMA"], header_format)
            worksheet.write(row_index, 9, ubicacion["TOTAL_DIFERENCIA"], header_format)
            row_index += 2

        # Fila final: Total Diferencia global
        worksheet.write(row_index, 0, "DIFERENCIAS TOTAL", header_format)
        worksheet.write(row_index, 9, datos_diferencias["DIFERENCIA_TOTAL"], header_format)

        workbook.close()
        output.seek(0)

        response = make_response(output.read())
        response.headers["Content-Disposition"] = "attachment; filename=diferencias.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return response
    finally:
        db.close()
    
def json_serial(obj):
    if isinstance(obj, datetime.datetime):
        return obj.isoformat()
    raise TypeError("Tipo no serializable")

def obtener_diferencias(db,id_planificacion,pais):
    """
    La tabla CAPTACION_FISICA tiene un campo ID_PLANIFICACION necesito considerar un condicion donde solo tome 
    las capturas de que corresponden al id_planificacion
    """
    try:
        utils = Utils()
        
        #Este query obtiene las captaciones fisicas agrupadas por ARTICULO, UBICACION, ALMACEN, LOTE, FECHA_EXPIRACION y ETIQUETA
        captaciones = db.query(
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            modelo.CaptacionFisica.FECHA_EXPIRACION,
            modelo.CaptacionFisica.ETIQUETA,
            modelo.CaptacionFisica.SERIE,
            modelo.CaptacionFisica.MODELO,
            func.round(func.sum(modelo.CaptacionFisica.CANTIDAD), 3).label("cantidad_captacion")
        ).filter(
            modelo.CaptacionFisica.ID_PLANIFICACION == id_planificacion,
            or_(
                modelo.CaptacionFisica.ETIQUETA.is_(None),
                modelo.CaptacionFisica.ETIQUETA != 'EN_TRANSITO'
            )
        ).group_by(
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            modelo.CaptacionFisica.FECHA_EXPIRACION,
            modelo.CaptacionFisica.ETIQUETA,
            modelo.CaptacionFisica.SERIE,
            modelo.CaptacionFisica.MODELO
        ).all()

        # Debug: imprimir las captaciones obtenidas
        for cap in captaciones:
            print("Captacion:", cap)

        filtros = utils.obtener_filtros(db,id_planificacion,pais)
        existencias = [dict(row._mapping) for row in utils.obtener_existencias_planificadas(db, filtros,id_planificacion)]
        print(f"existencias:")
        pprint.pprint(existencias)
        costos_por_ubicacion_y_almacen = utils.obtener_costos_por_ubicacion_y_almacen(db, filtros, id_planificacion)
        pprint.pprint(costos_por_ubicacion_y_almacen)
        existencias_dict = {}
        for e in existencias:
            clave = (
                e["ARTICULO"],
                e["UBICACION"],
                e["ALMACEN"],
                e["LOTE"],
                e["FECHA_EXPIRACION"],
                e.get("ETIQUETA"),
                e.get("SERIE"),
                e.get("MODELO")
            )
            e.setdefault("SERIE", None)
            e.setdefault("MODELO", None)
            existencias_dict[clave] = e
        print(f"Existencias encontradas:")
        pprint.pprint(existencias_dict)

        for cap in captaciones:
            if cap.ETIQUETA and cap.ETIQUETA.upper().startswith('EN_TRANSITO'):
                continue
            key = (cap.ARTICULO, cap.UBICACION, cap.ALMACEN, cap.LOTE, cap.FECHA_EXPIRACION, cap.ETIQUETA, cap.SERIE, cap.MODELO)
            existencia = existencias_dict.get(key)
            print(f"Procesando captación: {cap.ARTICULO} en {cap.UBICACION} - {cap.ALMACEN} - {cap.LOTE} - {cap.FECHA_EXPIRACION}")
            if existencia:
                existencia["cantidad_captacion"] = cap.cantidad_captacion
                existencia["diferencia"] = round(float(cap.cantidad_captacion) - float(existencia["cantidad_existencia"]), 3)
                existencia["ETIQUETA"] = cap.ETIQUETA
                existencia["costo_existencia"] = existencia["costo_existencia"]
                existencia["SERIE"] = cap.SERIE
                existencia["MODELO"] = cap.MODELO
            else:
                #TODO revisar buscar_costo_por_ubicacion_y_almacen
                print(f"Advertencia: No se encontró la existencia para la captación {cap.ARTICULO} en {cap.UBICACION} - {cap.ALMACEN} - {cap.LOTE} - {cap.FECHA_EXPIRACION}")
                existencias_dict[key] = {
                    "ARTICULO": cap.ARTICULO,
                    "DESCRIPCION": cap.DESCRIPCION,
                    "UBICACION": cap.UBICACION,
                    "ALMACEN": cap.ALMACEN,
                    "LOTE": cap.LOTE,
                    "FECHA_EXPIRACION": cap.FECHA_EXPIRACION,
                    "cantidad_captacion": cap.cantidad_captacion,
                    "cantidad_existencia": 0,
                    "costo_existencia": buscar_costo_por_ubicacion_y_almacen(cap.UBICACION, cap.ALMACEN, cap.ARTICULO, costos_por_ubicacion_y_almacen),
                    "diferencia": cap.cantidad_captacion,
                    "ETIQUETA": cap.ETIQUETA,
                    "SERIE": cap.SERIE,
                    "MODELO": cap.MODELO
                }

        for registro in existencias_dict.values():
            registro.setdefault("SERIE", None)
            registro.setdefault("MODELO", None)
            if registro.get("cantidad_captacion", 0) == 0:
                registro["diferencia"] = round((0 - registro.get("cantidad_existencia", 0)), 3)
                registro["cantidad_captacion"] = 0
                registro["cantidad_existencia"] = round(registro.get("cantidad_existencia", 0), 3)
                
        return list(existencias_dict.values())
    finally:
        db.close()

def buscar_costo_por_ubicacion_y_almacen(ubicacion, almacen, articulo, costos_por_ubicacion_y_almacen):
    for rec in costos_por_ubicacion_y_almacen:
        # rec is expected as a tuple: (UBICACION, ALMACEN, ARTICULO, costo_promedio)
        if rec[0] == ubicacion and rec[1] == almacen and rec[2] == articulo:
            return rec[3]
    return 0.0

def obtener_diccionario(resultados, comentarios_por_clave=None):
    """
    Este método crear un diccionario con la siguiente estructura:
    {
        "UBICACION": {
            "ARTICULO": {
                "CAPTACIONES": [
                    {
                        "ALMACEN": "",
                        "LOTE_NUEVO": "",
                        "FECHA_DE_EXP_NUEVO": "",
                        "LOTE": "",
                        "FECHA_EXPIRACION": "",
                        "FISICO": 0,
                        "SISTEMA": 0,
                        "DIFERENCIA": 0,
                        "RECONTEO": [],
                        "TOTAL_RECONTEO": 0
                    }
                ],
                "CAPTACIONES_NUEVAS": [],
                "DESCRIPCION": ""
            }
        }
    }
    """
    if comentarios_por_clave is None:
        comentarios_por_clave = {}

    data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for row in resultados:
        ubicacion = row["UBICACION"]
        articulo = row["ARTICULO"]
        descripcion = row["DESCRIPCION"]
        almacen = row["ALMACEN"]
        lote = row["LOTE"]
        fecha_expiracion = row["FECHA_EXPIRACION"]
        sistema = row["cantidad_existencia"]
        serie = row.get("SERIE") or None
        modelo = row.get("MODELO") or None
        #COMPARAR
        if "cantidad_captacion" not in row:
            row["cantidad_captacion"] = 0
            row["diferencia"] = 0 - sistema
        
        fisico = row["cantidad_captacion"]
        diferencia = row["diferencia"]
        if "ETIQUETA" not in row:
            row["ETIQUETA"] = None
        etiqueta = row["ETIQUETA"]
        etiqueta_cruda = row.get("ETIQUETA")

        if etiqueta and etiqueta.upper().startswith('EN_TRANSITO'):
            continue

        reconteo = "NO"
        if etiqueta:
            if etiqueta.startswith("RECONTEO"):
                reconteo = "SI"
                etiqueta = etiqueta.replace("RECONTEO_", "", 1)

        lote_nuevo = ""
        fecha_de_exp_nuevo = ""
        if etiqueta == "LOTE_NUEVO":
            lote_nuevo = "SI"
        elif etiqueta == "FECHA_NUEVA":
            fecha_de_exp_nuevo = "SI"
        elif etiqueta == "LOTE_FECHA_NUEVA":
            lote_nuevo = "SI"
            fecha_de_exp_nuevo = "SI"
        else:
            lote_nuevo = "NO"
            fecha_de_exp_nuevo = "NO"

        clave_comentario = (
            ubicacion,
            articulo,
            almacen,
            lote,
            fecha_expiracion,
            serie,
            modelo,
            etiqueta_cruda,
        )
        comentarios = comentarios_por_clave.get(clave_comentario, [])

        captacion = {
            "ALMACEN": almacen,
            "LOTE_NUEVO": lote_nuevo,
            "FECHA_DE_EXP_NUEVO": fecha_de_exp_nuevo,
            "LOTE": lote,
            "FECHA_EXPIRACION": fecha_expiracion,
            "FISICO": fisico,
            "SISTEMA": sistema,
            "DIFERENCIA": diferencia,
            "RECONTEO": [{"FISICO": fisico}],  # Ajusta según sea necesario
            "TOTAL_RECONTEO": fisico,  # Ajusta según sea necesario
            "RECONTEO": reconteo,
            "SERIE": serie,
            "MODELO": modelo,
            "TIENE_COMENTARIO": bool(comentarios),
            "COMENTARIOS_INLINE": [c.get("OBSERVACION", "") for c in comentarios if c.get("OBSERVACION")],
            "COMENTARIOS_TOTAL": len(comentarios),
        }

        if lote_nuevo == "NO" and fecha_de_exp_nuevo == "NO":
            data[ubicacion][articulo]['CAPTACIONES'].append(captacion)
        else:
            data[ubicacion][articulo]['CAPTACIONES_NUEVAS'].append(captacion)

        data[ubicacion][articulo]['DESCRIPCION'] = descripcion
    return data


def filtrar_diccionario_nuevos(diccionario):
    """Construye un nuevo diccionario solo con captaciones marcadas como nuevas."""
    filtrado = {}
    for ubicacion, articulos in diccionario.items():
        articulos_nuevos = {}
        for articulo, detalles in articulos.items():
            captaciones_nuevas = [deepcopy(captacion) for captacion in detalles.get('CAPTACIONES_NUEVAS', []) if captacion]
            if captaciones_nuevas:
                articulos_nuevos[articulo] = {
                    'DESCRIPCION': detalles['DESCRIPCION'],
                    'CAPTACIONES': [],
                    'CAPTACIONES_NUEVAS': captaciones_nuevas,
                }
        if articulos_nuevos:
            filtrado[ubicacion] = articulos_nuevos
    return filtrado

def obtener_diccionario_costo(resultados):
    """
    Este método crear un diccionario con la siguiente estructura:
    {
        "UBICACION": {
            "ARTICULO": {
                "CAPTACIONES": [
                    {
                        "ALMACEN": "",
                        "LOTE_NUEVO": "",
                        "FECHA_DE_EXP_NUEVO": "",
                        "LOTE": "",
                        "FECHA_EXPIRACION": "",
                        "FISICO": 0,
                        "SISTEMA": 0,
                        "DIFERENCIA": 0,
                        "RECONTEO": [],
                        "TOTAL_RECONTEO": 0
                    }
                ],
                "CAPTACIONES_NUEVAS": [],
                "DESCRIPCION": ""
            }
        }
    }
    """
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for row in resultados:
        ubicacion = row["UBICACION"]
        articulo = row["ARTICULO"]
        descripcion = row["DESCRIPCION"]
        almacen = row["ALMACEN"]
        lote = row["LOTE"]
        fecha_expiracion = row["FECHA_EXPIRACION"]

        #COMPARAR
        sistema = 0.00
        diferencia = 0.00
        sistema = 0.00
        if not float(row["costo_existencia"]) == 0:
            diferencia = round((float(row["diferencia"]) * float(row["costo_existencia"])), 3)
            fisico = round((float(row["cantidad_captacion"]) * float(row["costo_existencia"])), 3)
            sistema = round((float(row["cantidad_existencia"]) * float(row["costo_existencia"])), 3)

        if "ETIQUETA" not in row:
            row["ETIQUETA"] = None
        etiqueta = row["ETIQUETA"]

        reconteo = "NO"
        if etiqueta:
            if etiqueta.startswith("RECONTEO"):
                reconteo = "SI"
                etiqueta = etiqueta.replace("RECONTEO_", "", 1)

        lote_nuevo = ""
        fecha_de_exp_nuevo = ""
        if etiqueta == "LOTE_NUEVO":
            lote_nuevo = "SI"
        elif etiqueta == "FECHA_NUEVA":
            fecha_de_exp_nuevo = "SI"
        elif etiqueta == "LOTE_FECHA_NUEVA":
            lote_nuevo = "SI"
            fecha_de_exp_nuevo = "SI"
        else:
            lote_nuevo = "NO"
            fecha_de_exp_nuevo = "NO"

        captacion = {
            "ALMACEN": almacen,
            "LOTE_NUEVO": lote_nuevo,
            "FECHA_DE_EXP_NUEVO": fecha_de_exp_nuevo,
            "LOTE": lote,
            "FECHA_EXPIRACION": fecha_expiracion,
            "FISICO": fisico,
            "SISTEMA": sistema,
            "DIFERENCIA": diferencia,
            "RECONTEO": [{"FISICO": fisico}],  # Ajusta según sea necesario
            "TOTAL_RECONTEO": fisico,  # Ajusta según sea necesario
            "RECONTEO": reconteo,
            "SERIE": row.get("SERIE"),
            "MODELO": row.get("MODELO")
        }

        if lote_nuevo == "NO" and fecha_de_exp_nuevo == "NO":
            data[ubicacion][articulo]['CAPTACIONES'].append(captacion)
        else:
            data[ubicacion][articulo]['CAPTACIONES_NUEVAS'].append(captacion)

        data[ubicacion][articulo]['DESCRIPCION'] = descripcion
    return data


def captaciones_por_ubicacion(diccionario):
    """
    Este método toma el diccionario creado por el método anterior y lo transforma
    en una lista agrupada por ubicacion y dentro una lista de articulos y cada 
    articulo tiene una lista de CAPTACIONES y CAPTACIONES_NUEVAS
    """
    datos_diferencias = {
        "captaciones_por_ubicacion": []
    }

    #diccionario.items() retorna una lista de tuplas (clave, valor)
    #la clave es la ubicacion y el valor es otro diccionario 
    for ubicacion, articulos in diccionario.items():
        ubicacion_data = {
            "UBICACION": ubicacion,
            "ARTICULOS": [],
            "TOTAL_FISICO": 0,
            "TOTAL_SISTEMA": 0,
            "TOTAL_DIFERENCIA": 0
        }
        # ahora la clave es el articulo y el valor es otro diccionario con los detalles
        # de la captacion
        # y la captacion nueva
        for articulo, detalles in articulos.items():
            total_fisico = round(
            sum(c['FISICO'] for c in detalles.get('CAPTACIONES', [])) + 
            sum(c['FISICO'] for c in detalles.get('CAPTACIONES_NUEVAS', [])), 2)

            total_sistema = round(
                sum(c['SISTEMA'] for c in detalles.get('CAPTACIONES', [])) + 
                sum(c['SISTEMA'] for c in detalles.get('CAPTACIONES_NUEVAS', [])), 2)

            total_diferencia = round(
                sum(float(c['DIFERENCIA']) for c in detalles.get('CAPTACIONES', [])) + 
                sum(float(c['DIFERENCIA']) for c in detalles.get('CAPTACIONES_NUEVAS', [])), 2)
            
            articulo_data = {
                "ARTICULO": articulo,
                "DESCRIPCION": detalles['DESCRIPCION'],
                "CAPTACIONES": detalles.get('CAPTACIONES', []),
                "CAPTACIONES_NUEVAS": detalles.get('CAPTACIONES_NUEVAS', []),
                "TOTAL_FISICO": round(total_fisico, 3),
                "TOTAL_SISTEMA": round(total_sistema, 3),
                "TOTAL_DIFERENCIA": round(total_diferencia, 3)
            }
            ubicacion_data["ARTICULOS"].append(articulo_data)
            ubicacion_data["TOTAL_FISICO"] = round(ubicacion_data["TOTAL_FISICO"] + total_fisico, 3)
            ubicacion_data["TOTAL_SISTEMA"] = round(ubicacion_data["TOTAL_SISTEMA"] + total_sistema, 3)
            ubicacion_data["TOTAL_DIFERENCIA"] = round(ubicacion_data["TOTAL_DIFERENCIA"] + total_diferencia, 3)
        datos_diferencias["captaciones_por_ubicacion"].append(ubicacion_data)

    datos_diferencias["DIFERENCIA_TOTAL"] = round(sum(u["TOTAL_DIFERENCIA"] for u in datos_diferencias["captaciones_por_ubicacion"]), 3)
    return datos_diferencias


## -------------------------------------- ESTADOS -----------------------------------------
@bp.route("/estados/imagenes/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1','nivel_2','nivel_3','nivel_4'])
def reporte_imagenes(id_planificacion):
    try:
        db: Session = request.db
        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA

        # Consulta que retorna solo las captaciones que tienen imagen
        resultados = db.query(
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.CANTIDAD,
            modelo.CaptacionFisica.ESTADO,
            modelo.CaptacionFisica.IMAGEN
        ).filter(
            modelo.CaptacionFisica.IMAGEN.isnot(None),
            modelo.CaptacionFisica.IMAGEN != b'',
            modelo.CaptacionFisica.ID_PLANIFICACION == id_planificacion,
            or_(
                modelo.CaptacionFisica.ETIQUETA.is_(None),
                modelo.CaptacionFisica.ETIQUETA != 'EN_TRANSITO'
            )
        ).all()

        columnas = ["ARTICULO", "DESCRIPCION", "UBICACION", "ALMACEN", "CANTIDAD", "ESTADO", "IMAGEN"]
        resultados_modificados = []
        for row in resultados:
            row_dict = dict(zip(columnas, row))
            imagen_bytes = row_dict["IMAGEN"]
            row_dict["IMAGEN"] = "data:image/png;base64," + base64.b64encode(imagen_bytes).decode('utf-8')
            resultados_modificados.append(row_dict)
        
        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": date,
            "id_planificacion": id_planificacion,
            "report_catalog": report_catalog,
            "selected_report": "galeria",
            "resultados": resultados_modificados,
            "diferencias": None,
            "seccion_de_planificacion": False,
            "costo": False,
            "solo_nuevos": False,
            "sin_datos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()


@bp.route("/conteo/xls/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def conteo_xls(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)

        planificacion = _get_planificacion_actual(db, id_planificacion)
        filtros = _build_conteo_filter_options(db, utils, id_planificacion, pais, request.args, planificacion)
        datos_conteo = obtener_reporte_conteo(db, id_planificacion, pais, filtros["criteria"])

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Conteo")

        title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'align': 'center'})
        cell_format = workbook.add_format({'border': 1})
        number_format = workbook.add_format({'border': 1, 'num_format': '#,##0.000'})

        row_idx = 0
        worksheet.merge_range(row_idx, 0, row_idx, 10, "REPORTE DE CONTEO POR USUARIO", title_format)
        row_idx += 2

        headers = [
            "Usuario",
            "Ubicación",
            "Almacén",
            "Artículo",
            "Descripción",
            "Lote",
            "F. expiración",
            "Planificado",
            "Contado",
            "Diferencia",
            "Fecha captura",
        ]
        for col, header in enumerate(headers):
            worksheet.write(row_idx, col, header, header_format)
        row_idx += 1

        for fila in datos_conteo["rows"]:
            worksheet.write(row_idx, 0, fila.get("USUARIO") or "Sin captador", cell_format)
            worksheet.write(row_idx, 1, fila.get("UBICACION") or "", cell_format)
            worksheet.write(row_idx, 2, fila.get("ALMACEN") or "", cell_format)
            worksheet.write(row_idx, 3, fila.get("ARTICULO") or "", cell_format)
            worksheet.write(row_idx, 4, fila.get("DESCRIPCION") or "", cell_format)
            worksheet.write(row_idx, 5, fila.get("LOTE") or "", cell_format)

            fecha_exp = fila.get("FECHA_EXPIRACION")
            if isinstance(fecha_exp, datetime.datetime):
                worksheet.write(row_idx, 6, fecha_exp.strftime('%Y-%m-%d'), cell_format)
            elif isinstance(fecha_exp, datetime.date):
                worksheet.write(row_idx, 6, fecha_exp.strftime('%Y-%m-%d'), cell_format)
            else:
                worksheet.write(row_idx, 6, "", cell_format)

            worksheet.write_number(row_idx, 7, fila.get("CANTIDAD_EXISTENCIA") or 0, number_format)
            worksheet.write_number(row_idx, 8, fila.get("CANTIDAD_CAPTACION") or 0, number_format)
            worksheet.write_number(row_idx, 9, fila.get("DIFERENCIA") or 0, number_format)

            fecha_cap = fila.get("FECHA_CAPTURA")
            if isinstance(fecha_cap, datetime.datetime):
                worksheet.write(row_idx, 10, fecha_cap.strftime('%Y-%m-%d %H:%M'), cell_format)
            elif isinstance(fecha_cap, datetime.date):
                worksheet.write(row_idx, 10, fecha_cap.strftime('%Y-%m-%d'), cell_format)
            else:
                worksheet.write(row_idx, 10, "", cell_format)

            row_idx += 1

        row_idx += 1
        worksheet.write(row_idx, 6, "Totales", header_format)
        worksheet.write_number(row_idx, 7, datos_conteo["totales"]["plan"], number_format)
        worksheet.write_number(row_idx, 8, datos_conteo["totales"]["captado"], number_format)
        worksheet.write_number(row_idx, 9, datos_conteo["totales"]["diferencia"], number_format)

        workbook.close()
        output.seek(0)

        response = make_response(output.getvalue())
        filename = f"conteo_{id_planificacion}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    finally:
        db.close()


@bp.route("/consolidado/xls/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def consolidado_xls(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)

        planificacion = _get_planificacion_actual(db, id_planificacion)
        resultados = obtener_diferencias(db, id_planificacion, pais)
        datos_consolidado = obtener_diccionario_consolidado(resultados)

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Consolidado")

        title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'align': 'center'})
        cell_format = workbook.add_format({'border': 1})
        number_format = workbook.add_format({'border': 1, 'num_format': '#,##0.000'})

        row_idx = 0
        worksheet.merge_range(row_idx, 0, row_idx, 13, "REPORTE CONSOLIDADO CANTIDADES / COSTO", title_format)
        row_idx += 2

        headers = [
            "Ubicación",
            "Artículo",
            "Descripción",
            "Almacén",
            "Lote",
            "Serie",
            "Modelo",
            "F. expiración",
            "Físico",
            "Sistema",
            "Diferencia",
            "Físico (Costo)",
            "Sistema (Costo)",
            "Diferencia (Costo)",
        ]
        for col, header in enumerate(headers):
            worksheet.write(row_idx, col, header, header_format)
        row_idx += 1

        for ubicacion in datos_consolidado.get("captaciones_por_ubicacion", []):
            nombre_ubicacion = ubicacion.get("UBICACION") or ""
            for articulo in ubicacion.get("ARTICULOS", []):
                codigo_articulo = articulo.get("ARTICULO") or ""
                descripcion = articulo.get("DESCRIPCION") or ""
                for captacion in articulo.get("CAPTACIONES", []):
                    worksheet.write(row_idx, 0, nombre_ubicacion, cell_format)
                    worksheet.write(row_idx, 1, codigo_articulo, cell_format)
                    worksheet.write(row_idx, 2, descripcion, cell_format)
                    worksheet.write(row_idx, 3, captacion.get("ALMACEN") or "", cell_format)
                    worksheet.write(row_idx, 4, captacion.get("LOTE") or "", cell_format)
                    worksheet.write(row_idx, 5, captacion.get("SERIE") or "", cell_format)
                    worksheet.write(row_idx, 6, captacion.get("MODELO") or "", cell_format)

                    fecha_exp = captacion.get("FECHA_EXPIRACION")
                    if isinstance(fecha_exp, datetime.datetime):
                        worksheet.write(row_idx, 7, fecha_exp.strftime('%Y-%m-%d'), cell_format)
                    elif isinstance(fecha_exp, datetime.date):
                        worksheet.write(row_idx, 7, fecha_exp.strftime('%Y-%m-%d'), cell_format)
                    else:
                        worksheet.write(row_idx, 7, captacion.get("FECHA_EXPIRACION_STR") or "", cell_format)

                    worksheet.write_number(row_idx, 8, captacion.get("FISICO") or 0, number_format)
                    worksheet.write_number(row_idx, 9, captacion.get("SISTEMA") or 0, number_format)
                    worksheet.write_number(row_idx, 10, captacion.get("DIFERENCIA") or 0, number_format)
                    worksheet.write_number(row_idx, 11, captacion.get("FISICO_COSTO") or 0, number_format)
                    worksheet.write_number(row_idx, 12, captacion.get("SISTEMA_COSTO") or 0, number_format)
                    worksheet.write_number(row_idx, 13, captacion.get("DIFERENCIA_COSTO") or 0, number_format)
                    row_idx += 1

        row_idx += 1
        worksheet.write(row_idx, 9, "TOTAL DIF.", header_format)
        worksheet.write_number(row_idx, 10, datos_consolidado.get("DIFERENCIA_TOTAL", 0), number_format)
        worksheet.write(row_idx, 12, "TOTAL DIF. COSTO", header_format)
        worksheet.write_number(row_idx, 13, datos_consolidado.get("DIFERENCIA_TOTAL_COSTO", 0), number_format)

        workbook.close()
        output.seek(0)

        response = make_response(output.getvalue())
        filename = f"consolidado_{id_planificacion}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    finally:
        db.close()

@bp.route("/estados/<id_planificacion>",methods=['GET'])
@requiere_login
@role_required(['nivel_1','nivel_2','nivel_3','nivel_4'])
def reporte_estados(id_planificacion):
    try:
        db: Session = request.db
        planificacion = _get_planificacion_actual(db, id_planificacion)
        date = planificacion.FECHA
        # Se obtiene la lista de estados
        datos_estados = obtener_datos_estados(db,id_planificacion)
        report_catalog = build_report_catalog(id_planificacion)
        context = {
            "usuario": g.user.usuario,
            "date": date,
            "id_planificacion": id_planificacion,
            "datos_estados": datos_estados,
            "report_catalog": report_catalog,
            "selected_report": "estados",
            "diferencias": None,
            "seccion_de_planificacion": False,
            "costo": False,
            "solo_nuevos": False,
            "sin_datos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/diferencias.html", **context)
    finally:
        db.close()

@bp.route("/estados/xls/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def generar_reporte_estados_xls(id_planificacion):
    try:
        db = SessionLocal()
        _get_planificacion_actual(db, id_planificacion)
        datos_estados = obtener_datos_estados(db, id_planificacion)

        # Crear un archivo Excel en memoria
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte Estados"

        # Estilo para encabezados
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")

        # Encabezado principal
        sheet.merge_cells("A1:H1")
        sheet["A1"] = "REPORTE POR ESTADOS DEL INVENTARIO FÍSICO"
        sheet["A1"].font = header_font
        sheet["A1"].alignment = center_alignment

        # Fecha
        sheet.merge_cells("A3:H3")
        sheet["A3"] = f"FECHA: {datetime.datetime.now()}"
        sheet["A3"].font = header_font

        # Procesar datos
        row = 5
        for ubicacion in datos_estados["captaciones_por_ubicacion"]:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            sheet.cell(row=row, column=1).value = f"UBICACIÓN: {ubicacion['UBICACION']}"
            sheet.cell(row=row, column=1).font = header_font
            row += 1

            headers = ["ARTICULO", "DESCRIPCION", "ALMACEN", "LOTE", "F. EXPIRACION", "FISICO", "ESTADO"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=row, column=col).value = header
                sheet.cell(row=row, column=col).font = header_font
                sheet.cell(row=row, column=col).alignment = center_alignment
            row += 1

            for articulo in ubicacion["ARTICULOS"]:
                for captacion in articulo["CAPTACIONES"]:
                    sheet.cell(row=row, column=1).value = captacion["ARTICULO"]
                    sheet.cell(row=row, column=2).value = captacion["DESCRIPCION"]
                    sheet.cell(row=row, column=3).value = captacion["ALMACEN"]
                    sheet.cell(row=row, column=4).value = captacion["LOTE"]
                    sheet.cell(row=row, column=5).value = captacion["FECHA_DE_EXPIRACION"]
                    sheet.cell(row=row, column=6).value = captacion["FISICO"]
                    sheet.cell(row=row, column=7).value = captacion["ESTADO"]
                    row += 1

            sheet.cell(row=row, column=5).value = "TOTAL"
            sheet.cell(row=row, column=6).value = ubicacion["TOTAL_UBICACION"]
            sheet.cell(row=row, column=5).font = header_font
            sheet.cell(row=row, column=6).font = header_font
            row += 2

        sheet.cell(row=row, column=5).value = "TOTAL GLOBAL"
        sheet.cell(row=row, column=6).value = datos_estados["TOTAL_GLOBAL"]
        sheet.cell(row=row, column=5).font = header_font
        sheet.cell(row=row, column=6).font = header_font

        workbook.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=reporte_estados.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    finally:
        db.close()

def obtener_datos_estados(db, id_planificacion):
    resultados = obtener_capturas_por_estado(db, id_planificacion)
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for result in resultados:
        ubicacion = result.UBICACION
        almacen = result.ALMACEN
        lote = result.LOTE or ''  # Reemplazar None por ''
        fecha_expiracion = result.FECHA_EXPIRACION   # Convertir datetime a cadena
        estado = result.ESTADO
        total_cantidad = result.total_cantidad
        articulo = result.ARTICULO
        descripcion = result.DESCRIPCION

        data[ubicacion][lote][estado].append({
            "UBICACION": ubicacion,
            "ALMACEN": almacen,
            "LOTE": lote,
            "FECHA_DE_EXPIRACION": fecha_expiracion,  # Fecha como cadena
            "FISICO": total_cantidad,
            "ESTADO": estado,
            "ARTICULO": articulo,
            "DESCRIPCION": descripcion
        })

    datos_estados = {
        "captaciones_por_ubicacion": []
    }

    total_global = 0

    for ubicacion, lotes in data.items():
        total_ubicacion = 0
        articulos = []
        for lote, estados in lotes.items():
            for estado, captaciones in estados.items():
                total_fisico = sum(captacion["FISICO"] for captacion in captaciones)
                total_ubicacion += total_fisico
                articulos.append({
                    "ARTICULO": captaciones[0]["ARTICULO"],
                    "DESCRIPCION": captaciones[0]["DESCRIPCION"],
                    "CAPTACIONES": captaciones,
                    "TOTAL_FISICO": total_fisico
                })
        datos_estados["captaciones_por_ubicacion"].append({
            "UBICACION": ubicacion,
            "ARTICULOS": articulos,
            "TOTAL_UBICACION": total_ubicacion
        })
        total_global += total_ubicacion

    datos_estados["TOTAL_GLOBAL"] = total_global
    return datos_estados

def obtener_capturas_por_estado(db, id_planificacion=None):
    try:
        query = db.query(
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            func.to_char(modelo.CaptacionFisica.FECHA_EXPIRACION, 'DD/MM/YYYY').label('FECHA_EXPIRACION'),  # Formatear la fecha
            modelo.CaptacionFisica.ESTADO,
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,  # Agregar UBICACION
            func.sum(modelo.CaptacionFisica.CANTIDAD).label('total_cantidad')
        ).filter(
            modelo.CaptacionFisica.ESTADO.isnot(None),
            modelo.CaptacionFisica.ESTADO != '',
            modelo.CaptacionFisica.ID_PLANIFICACION == id_planificacion,
            or_(
                modelo.CaptacionFisica.ETIQUETA.is_(None),
                modelo.CaptacionFisica.ETIQUETA != 'EN_TRANSITO'
            )
        ).group_by(
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            modelo.CaptacionFisica.FECHA_EXPIRACION,
            modelo.CaptacionFisica.ESTADO,
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION  # Agregar UBICACION al group_by
        )

        resultados = query.all()
        return resultados
    finally:
        db.close()


def obtener_datos_transito(db, id_planificacion, pais):
    """
    Obtiene todas las captaciones marcadas como EN_TRANSITO con sus costos.
    Retorna estructura agrupada por ubicación con flags para estados e imágenes.
    """
    try:
        utils = Utils()
        
        # Consultar captaciones EN_TRANSITO
        captaciones = db.query(
            modelo.CaptacionFisica.ARTICULO,
            modelo.CaptacionFisica.DESCRIPCION,
            modelo.CaptacionFisica.UBICACION,
            modelo.CaptacionFisica.ALMACEN,
            modelo.CaptacionFisica.LOTE,
            modelo.CaptacionFisica.FECHA_EXPIRACION,
            modelo.CaptacionFisica.CANTIDAD,
            modelo.CaptacionFisica.ESTADO,
            modelo.CaptacionFisica.OBSERVACION,
            modelo.CaptacionFisica.USUARIO,
            modelo.CaptacionFisica.FECHA,
            modelo.CaptacionFisica.IMAGEN,
            modelo.CaptacionFisica.ETIQUETA
        ).filter(
            modelo.CaptacionFisica.ID_PLANIFICACION == id_planificacion,
            modelo.CaptacionFisica.ETIQUETA.like('EN_TRANSITO%')
        ).order_by(
            modelo.CaptacionFisica.UBICACION.asc(),
            modelo.CaptacionFisica.FECHA.desc()
        ).all()
        
        # Obtener costos por ubicación/almacén/artículo
        filtros = utils.obtener_filtros(db, id_planificacion, pais)
        costos_por_ubicacion_y_almacen = utils.obtener_costos_por_ubicacion_y_almacen(db, filtros, id_planificacion)
        
        # Detectar si hay estados o imágenes
        tiene_estados = any(cap.ESTADO is not None and cap.ESTADO != '' for cap in captaciones)
        tiene_imagenes = any(cap.IMAGEN is not None and cap.IMAGEN != b'' for cap in captaciones)
        
        # Agrupar por ubicación
        ubicaciones_dict = defaultdict(list)
        
        for cap in captaciones:
            # Buscar costo unitario
            costo_unitario = float(buscar_costo_por_ubicacion_y_almacen(
                cap.UBICACION, 
                cap.ALMACEN, 
                cap.ARTICULO, 
                costos_por_ubicacion_y_almacen
            ))
            
            cantidad = float(cap.CANTIDAD or 0)
            costo_total = round(cantidad * costo_unitario, 3)
            
            # Formatear fechas
            fecha_exp_str = ""
            if cap.FECHA_EXPIRACION:
                if isinstance(cap.FECHA_EXPIRACION, datetime.datetime):
                    fecha_exp_str = cap.FECHA_EXPIRACION.strftime('%d/%m/%Y')
                elif isinstance(cap.FECHA_EXPIRACION, datetime.date):
                    fecha_exp_str = cap.FECHA_EXPIRACION.strftime('%d/%m/%Y')
            
            fecha_captura_str = ""
            if cap.FECHA:
                if isinstance(cap.FECHA, datetime.datetime):
                    fecha_captura_str = cap.FECHA.strftime('%d/%m/%Y %H:%M')
                elif isinstance(cap.FECHA, datetime.date):
                    fecha_captura_str = cap.FECHA.strftime('%d/%m/%Y')
            
            # Convertir imagen a base64 si existe
            imagen_base64 = None
            if cap.IMAGEN and cap.IMAGEN != b'':
                imagen_base64 = "data:image/png;base64," + base64.b64encode(cap.IMAGEN).decode('utf-8')
            
            ubicaciones_dict[cap.UBICACION or "SIN_UBICACION"].append({
                "ARTICULO": cap.ARTICULO,
                "DESCRIPCION": cap.DESCRIPCION,
                "UBICACION": cap.UBICACION,
                "ALMACEN": cap.ALMACEN,
                "LOTE": cap.LOTE,
                "FECHA_EXPIRACION": fecha_exp_str,
                "CANTIDAD": cantidad,
                "COSTO_UNITARIO": costo_unitario,
                "COSTO_TOTAL": costo_total,
                "ESTADO": cap.ESTADO,
                "OBSERVACION": cap.OBSERVACION,
                "USUARIO": cap.USUARIO,
                "FECHA": fecha_captura_str,
                "IMAGEN": imagen_base64,
                "TIENE_IMAGEN": imagen_base64 is not None
            })
        
        # Construir estructura final
        captaciones_por_ubicacion = []
        total_global_cantidad = 0.0
        total_global_costo = 0.0
        
        for ubicacion, captaciones_list in sorted(ubicaciones_dict.items()):
            total_ubicacion_cantidad = sum(c["CANTIDAD"] for c in captaciones_list)
            total_ubicacion_costo = sum(c["COSTO_TOTAL"] for c in captaciones_list)
            
            captaciones_por_ubicacion.append({
                "UBICACION": ubicacion,
                "CAPTACIONES": captaciones_list,
                "TOTAL_CANTIDAD": round(total_ubicacion_cantidad, 3),
                "TOTAL_COSTO": round(total_ubicacion_costo, 3)
            })
            
            total_global_cantidad += total_ubicacion_cantidad
            total_global_costo += total_ubicacion_costo
        
        return {
            "captaciones_por_ubicacion": captaciones_por_ubicacion,
            "TOTAL_GLOBAL_CANTIDAD": round(total_global_cantidad, 3),
            "TOTAL_GLOBAL_COSTO": round(total_global_costo, 3),
            "tiene_estados": tiene_estados,
            "tiene_imagenes": tiene_imagenes
        }
    
    except Exception as e:
        print(f"Error en obtener_datos_transito: {e}")
        import traceback
        traceback.print_exc()
        raise


@bp.route("/transito/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def reporte_transito(id_planificacion):
    try:
        db: Session = request.db
        utils = Utils()
        pais = utils.obtener_pais(db, g)
        
        planificacion = _get_planificacion_actual(db, id_planificacion)
        datos_transito = obtener_datos_transito(db, id_planificacion, pais)
        report_catalog = build_report_catalog(id_planificacion)
        
        context = {
            "usuario": g.user.usuario,
            "date": planificacion.FECHA,
            "id_planificacion": id_planificacion,
            "datos_transito": datos_transito,
            "report_catalog": report_catalog,
            "selected_report": "transito",
            "diferencias": None,
            "seccion_de_planificacion": False,
            "costo": False,
            "solo_nuevos": False,
            "sin_datos": False,
        }
        context.update(_build_planificacion_context(planificacion))
        return render_template("reporte/transito.html", **context)
    finally:
        db.close()


@bp.route("/transito/xls/<id_planificacion>", methods=['GET'])
@requiere_login
@role_required(['nivel_1', 'nivel_2', 'nivel_3', 'nivel_4'])
def generar_reporte_transito_xls(id_planificacion):
    try:
        db = SessionLocal()
        utils = Utils()
        pais = utils.obtener_pais(db, g)
        
        _get_planificacion_actual(db, id_planificacion)
        datos_transito = obtener_datos_transito(db, id_planificacion, pais)

        # Crear archivo Excel en memoria
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Artículos en Tránsito"

        # Estilos
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")

        # Encabezado principal
        sheet.merge_cells("A1:M1")
        sheet["A1"] = "REPORTE DE ARTÍCULOS EN TRÁNSITO"
        sheet["A1"].font = header_font
        sheet["A1"].alignment = center_alignment

        # Fecha
        sheet.merge_cells("A3:M3")
        sheet["A3"] = f"FECHA: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        sheet["A3"].font = header_font

        # Procesar datos
        row = 5
        for ubicacion_data in datos_transito["captaciones_por_ubicacion"]:
            # Encabezado de ubicación
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            sheet.cell(row=row, column=1).value = f"UBICACIÓN: {ubicacion_data['UBICACION']}"
            sheet.cell(row=row, column=1).font = header_font
            row += 1

            # Encabezados de columnas
            headers = ["ARTICULO", "DESCRIPCION", "UBICACION", "ALMACEN", "LOTE", "F.EXP", 
                      "CANTIDAD", "COSTO UNIT.", "COSTO TOTAL", "OBSERVACION", "CAPTADOR", "FECHA", "TIENE IMAGEN", "ESTADO"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=row, column=col).value = header
                sheet.cell(row=row, column=col).font = header_font
                sheet.cell(row=row, column=col).alignment = center_alignment
            row += 1

            # Datos de captaciones
            for captacion in ubicacion_data["CAPTACIONES"]:
                sheet.cell(row=row, column=1).value = captacion["ARTICULO"]
                sheet.cell(row=row, column=2).value = captacion["DESCRIPCION"]
                sheet.cell(row=row, column=3).value = captacion["UBICACION"]
                sheet.cell(row=row, column=4).value = captacion["ALMACEN"]
                sheet.cell(row=row, column=5).value = captacion["LOTE"]
                sheet.cell(row=row, column=6).value = captacion["FECHA_EXPIRACION"]
                sheet.cell(row=row, column=7).value = captacion["CANTIDAD"]
                sheet.cell(row=row, column=8).value = captacion["COSTO_UNITARIO"]
                sheet.cell(row=row, column=9).value = captacion["COSTO_TOTAL"]
                sheet.cell(row=row, column=10).value = captacion["OBSERVACION"] or ""
                sheet.cell(row=row, column=11).value = captacion["USUARIO"] or ""
                sheet.cell(row=row, column=12).value = captacion["FECHA"]
                sheet.cell(row=row, column=13).value = "SÍ" if captacion["TIENE_IMAGEN"] else "NO"
                sheet.cell(row=row, column=14).value = captacion["ESTADO"] or ""
                row += 1

            # Total por ubicación
            sheet.cell(row=row, column=6).value = "TOTAL UBICACIÓN"
            sheet.cell(row=row, column=7).value = ubicacion_data["TOTAL_CANTIDAD"]
            sheet.cell(row=row, column=9).value = ubicacion_data["TOTAL_COSTO"]
            sheet.cell(row=row, column=6).font = header_font
            sheet.cell(row=row, column=7).font = header_font
            sheet.cell(row=row, column=9).font = header_font
            row += 2

        # Total global
        sheet.cell(row=row, column=6).value = "TOTAL GLOBAL"
        sheet.cell(row=row, column=7).value = datos_transito["TOTAL_GLOBAL_CANTIDAD"]
        sheet.cell(row=row, column=9).value = datos_transito["TOTAL_GLOBAL_COSTO"]
        sheet.cell(row=row, column=6).font = header_font
        sheet.cell(row=row, column=7).font = header_font
        sheet.cell(row=row, column=9).font = header_font

        workbook.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=reporte_transito.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    finally:
        db.close()
